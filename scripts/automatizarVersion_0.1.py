import pyautogui
import time
import pandas as pd
import openpyxl
import win32com.client as win32
import pyautogui
import time
import win32gui, win32con
import routes as gv
# -------------------------------
# CONFIGURACIÓN
# -------------------------------
# Rutas
ruta_maestro = gv.RUTA_OUTPUT/ "LibroControl.xlsx"
# Coordenadas de los clicks (ejemplo, medir con mouseInfo)
coord_abajo=(362,425)
coord_arriba=(362,388)
coord_banco = (393, 363)
coord_obtener = (76, 409)
coord_desmarcar=(155,387)
coord_fecha=(155,423)
coord_ultima=(155,385)
coord_penultima=(156,404)

#Funciones auxialiares
def activar_ventana_al_frente(hwnd, timeout=5):
    start = time.time()
    while time.time() - start < timeout:
        if hwnd:
            # 💡 CAMBIO CLAVE: Usamos SW_MAXIMIZE en lugar de SW_RESTORE
            # SW_MAXIMIZE (3) maximiza la ventana.
            # SW_RESTORE (9) la restaura a su tamaño anterior (si estaba minimizada).
            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE) 
            
            # Traer al frente
            win32gui.SetForegroundWindow(hwnd) 
            break
        time.sleep(0.2)
    else:
        print(f"No se pudo activar/maximizar la ventana con hwnd {hwnd}")
def abrir_excel(ruta):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(ruta)
    excel.Visible = True 

    # 🌟 PASO CRUCIAL 1: Dale un momento a Windows para que dibuje la ventana.
    time.sleep(0.1) 
    
    # 🌟 PASO CRUCIAL 2: Obtener el Handle (hwnd) del objeto Excel.Application
    # La propiedad Hwnd está disponible en el objeto Application.
    hwnd_excel = excel.Hwnd
    
    # 🌟 PASO CRUCIAL 3: Llamar a la función con el Handle, no con el nombre.
    activar_ventana_al_frente(hwnd_excel)
    
    wb.Sheets(1).Activate()
    return excel, wb
def keepClick(t):
    pyautogui.mouseDown()
    time.sleep(t)
    pyautogui.mouseUp()
def moveToClickAndWait(x,y,t,tc):
    pyautogui.moveTo(x, y, t)
    keepClick(tc)
    time.sleep(t) 
def marcarSiguienteFecha(n):
    for i in range(n):
        moveToClickAndWait(coord_abajo[0],coord_abajo[1],0,0)
    moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
    moveToClickAndWait(coord_arriba[0],coord_arriba[1],0,0)
    moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
    moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)


def limpieza_datos_por_archivo(rutaArchivoParaLimpiar,NumerodeBimestresEnArchivo):
    excel, wb_cnbv = abrir_excel(rutaArchivoParaLimpiar)
    ws_cnbv = wb_cnbv.Sheets(1)


    def traer_tabla(origen):

        #se seleciona B9 para poder seleccioanr toda la tabla ya que ahi inicia y esto no cambia
        data = origen.Range("B9").CurrentRegion.Value
        df = pd.DataFrame(data[1:], columns=data[0])
        #La siguiente linea seleciona el origen que va a tomar el nombre de los archivos aqui debes selecioanr una celda en la que siempre haya la fecha de la consulta
        #para los arhcivos de uan tablas es C9 y los demas es c10
        nombre_extra = str(int(origen.Range("C10").Value))
        # Crear el nombre del archivo
        preparar_nombre=gv.limpiar_texto(str(rutaArchivoParaLimpiar.stem))
        nueva_carpeta=gv.RUTA_OUTPUT / preparar_nombre
        #crea la carpeta destino
        nueva_carpeta.mkdir(parents=True, exist_ok=True)
        nombre_archivo=f"{preparar_nombre}_{nombre_extra}.xlsx"
        ruta_final=nueva_carpeta/nombre_archivo
        # -------------------------------
        # GUARDAR EN LIBRO MAESTRO
        # -------------------------------
        try:
            wb_maestro = openpyxl.load_workbook(ruta_maestro)
            ws_maestro = wb_maestro.active
        except FileNotFoundError:
            wb_maestro = openpyxl.Workbook()
            ws_maestro = wb_maestro.active

        # Escribir datos en la hoja
        for r_idx, row in enumerate(df.values, 2):  # comienza en fila 2 para encabezado
            for c_idx, value in enumerate(row, 1):
                ws_maestro.cell(row=r_idx, column=c_idx, value=value)

        # Encabezados
        for c_idx, header in enumerate(df.columns, 1):
            ws_maestro.cell(row=1, column=c_idx, value=header)

        # Guardar libro maestro
        #LibroControl
        wb_maestro.save(gv.fix_long_path(str(ruta_final)))

        # -------------------------------
        # CERRAR LIBRO CNBV
        # -------------------------------
        activar_ventana_al_frente(excel.Hwnd)

        print("Tabla copiada correctamente al libro maestro.")


    max_iteraciones = NumerodeBimestresEnArchivo
    contador = 0

    while contador < max_iteraciones:
            valor_actual = ws_cnbv.Range("D37").Value
            time.sleep(0.5)
                # Aquí llamas a tu función de clicks, por ejemplo:
            if contador==0:
                print("Celda vacía, ejecutando clicks...")  
                time.sleep(0.5)
                moveToClickAndWait(coord_desmarcar[0],coord_desmarcar[1],0,0)
                #algunos archivos de la cnbv no es necesario marcar y desmarcar la casilal de bancos para selccioanr todos basta con un solo marcado
                #agregar o quitar la siguiente linea segun sea el caso
                moveToClickAndWait(coord_banco[0],coord_banco[1],0,0)
                moveToClickAndWait(coord_banco[0],coord_banco[1],2,0)
                moveToClickAndWait(coord_abajo[0],coord_abajo[1],0,6)
                moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                
            elif contador<=max_iteraciones-3:
                print(f"Celda cambió: {valor_actual}")
                # Continuar con siguiente acción o rango
                traer_tabla(ws_cnbv)
                marcarSiguienteFecha(20-contador)
            else:
                traer_tabla(ws_cnbv)
                moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
                moveToClickAndWait(coord_penultima[0],coord_penultima[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                traer_tabla(ws_cnbv)
                moveToClickAndWait(coord_penultima[0],coord_penultima[1],0,0)
                moveToClickAndWait(coord_ultima[0],coord_ultima[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                traer_tabla(ws_cnbv)
                break
            contador += 1
    
    wb_cnbv.Close(SaveChanges=False)        
    excel.Quit()
x=gv.ARCHIVOS_DATOSTARJETASCREDITO

#sigue el 17
#print(x[10])
#print(len(x))
limpieza_datos_por_archivo(x[10],85)
