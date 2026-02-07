import pandas as pd
import openpyxl
import win32com.client as win32
import pyautogui
import time
import win32gui, win32con
import routes as gv
from pathlib import Path
import pywintypes
import re
from typing import Any
import config as con

pyautogui.FAILSAFE = True
rutasInstaciadas=con.Paths()

# -------------------------------
# CONFIGURACIÓN
# -------------------------------
# Libro maestro donde se consolidan todos los datos
ruta_maestro = gv.RUTA_OUTPUT/ "LibroControl.xlsx"
# Coordenadas de los clicks
CONFIG_COORDS = {
    "abajo": (362, 425),
    "arriba": (362, 388),
    "banco": (393, 363),
    "obtener": (76, 409),
    "fecha_top": (155, 387),
    "fecha_bottom": (155, 423),
    "penultima": (156, 404),
    "tiempo_espera_consulta": 1,
    "tiempo_espera_entre_clicks":0.8
}

coord_abajo=(362,425)
coord_arriba=(362,388)
coord_banco = (393, 363)
coord_obtener = (76, 409)
coord_desmarcar=(155,387)
coord_fecha=(155,423)
coord_ultima=(155,385)
coord_penultima=(156,404)

def busqueda_patron_para_nombrar_carpetas(exp:str)-> str:
        # Explicación del patrón:
    # ^[^_]+  -> Salta el primer bloque (040)
    # _[^_]+  -> Salta el segundo bloque (12e)
    # _([^_]+) -> Captura lo que hay en el tercer bloque hasta el siguiente guion bajo
    patron = con.regexp_patrones.ARCHIVOSCNBV

    resultado = re.search(patron, exp)

    if resultado:
        serie_nombre = resultado.group()
        print(f"Serie extraída: {serie_nombre}") 
    return serie_nombre


#Funciones auxialiares
def activar_ventana_al_frente(hwnd, timeout=5)-> None:
    start = time.time()
    while time.time() - start < timeout:
        if hwnd:
            # Maximizar ventana
            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)             
            # Traer al frente
            win32gui.SetForegroundWindow(hwnd) 
            break
        time.sleep(0.2)
    else:
        print(f"No se pudo activar/maximizar la ventana con hwnd {hwnd}")

def abrir_excel(ruta, reintentos=3) -> tuple:
    # Usar EnsureDispatch es genial porque asegura que las constantes de Excel se carguen
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    
    try:
        wb = excel.Workbooks.Open(ruta)
        excel.Visible = True 
        
        # 🌟 PASO CRUCIAL 1: Dale un momento a Windows para que "pinte" el proceso
        time.sleep(0.5) # Subí un poco a 0.5 porque a veces 0.1 es muy rápido para el Render
        
        hwnd_excel = excel.Hwnd
        
        for i in range(reintentos):
            try:
                # Intentamos activar la ventana físicamente
                activar_ventana_al_frente(hwnd_excel)
                # Intentamos activar la hoja lógicamente
                wb.Sheets(1).Activate()
                
                print(f"Excel listo y al frente al intento {i+1}")
                return excel, wb
                
            except (pywintypes.error, Exception) as e:
                print(f"Intento {i+1} fallido: {e}")
                if i < reintentos - 1:
                    time.sleep(1)
                else:
                    print("Se agotaron los reintentos.")
                    # Si falla, cerramos para no dejar procesos "zombie" en el Administrador de Tareas
                    wb.Close(False)
                    excel.Quit()
                    raise Exception("No se pudo poner Excel en primer plano.")
                    
    except Exception as e:
        print(f"Error crítico al abrir el archivo: {e}")
        excel.Quit()
        return excel, wb

def keepClick(t)-> None:
    pyautogui.mouseDown()
    time.sleep(t)
    pyautogui.mouseUp()
def moveToClickAndWait(x:int,y:int,t:float,tc:float)-> None:
    pyautogui.moveTo(x, y, t)
    keepClick(tc)
    time.sleep(t) 
def marcarSiguienteFecha(n: int)-> None:
    for i in range(n):
        moveToClickAndWait(coord_abajo[0],coord_abajo[1],0,0)
    moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
    moveToClickAndWait(coord_arriba[0],coord_arriba[1],0,0)
    moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
    moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)




def limpieza_datos_por_archivo(rutaArchivoParaLimpiar,NumerodeBimestresEnArchivo)-> None:

    excel, wb_cnbv = abrir_excel(rutaArchivoParaLimpiar)
    ws_cnbv = wb_cnbv.Sheets(1)


    def traer_tabla(origen):

        #se seleciona B9 para poder seleccioanr toda la tabla ya que ahi inicia y esto no cambia
        data = origen.Range("B9").CurrentRegion.Value
        df = pd.DataFrame(data[1:], columns=data[0])
        #La siguiente linea seleciona el origen que va a tomar el nombre de los archivos aqui debes selecioanr una celda en la que siempre haya la fecha de la consulta
        #para los arhcivos de uan tablas es C9 y los demas es c10
        valor_c9 = origen.Range("C9").Value
        valor_c10 = origen.Range("C10").Value

        if valor_c9 is int:

            buscar_fecha = valor_c9
        else:
            buscar_fecha=valor_c10
        try:
            nombre_extra=str(buscar_fecha)
        except (ValueError, TypeError):
            print("Error: No se encontró una fecha válida en C9 ni en C10")
            nombre_extra = "FECHA_DESCONOCIDA"
        archivo_origen=busqueda_patron_para_nombrar_carpetas(rutaArchivoParaLimpiar)
        # Crear el nombre del archivo
        nueva_carpeta=gv.RUTA_OUTPUT / archivo_origen
        #crea la carpeta destino
        nueva_carpeta.mkdir(parents=True, exist_ok=True)
        nombre_archivo=f"{archivo_origen}_{nombre_extra}.xlsx"
        ruta_final=nueva_carpeta/nombre_archivo
        # -------------------------------
        # GUARDAR EN LIBRO MAESTRO
        # -------------------------------
        try:
            wb_maestro = openpyxl.load_workbook(ruta_maestro)
            ws_maestro = wb_maestro.active
        except FileNotFoundError:
            wb_maestro = openpyxl.Workbook()
            ws_maestro = wb_maestro.active

        # Escribir datos en la hoja
        for r_idx, row in enumerate(df.values, 2):  # comienza en fila 2 para encabezado
            for c_idx, value in enumerate(row, 1):
                ws_maestro.cell(row=r_idx, column=c_idx, value=value)

        # Encabezados
        for c_idx, header in enumerate(df.columns, 1):
            ws_maestro.cell(row=1, column=c_idx, value=header)

        # Guardar libro maestro
        #LibroControl
        wb_maestro.save(gv.fix_long_path(str(ruta_final)))

        # -------------------------------
        # CERRAR LIBRO CNBV
        # -------------------------------
        activar_ventana_al_frente(excel.Hwnd)

        print("Tabla copiada correctamente al libro maestro.")


    max_iteraciones = NumerodeBimestresEnArchivo
    contador = 0

    while contador < max_iteraciones:
            valor_actual = ws_cnbv.Range("D37").Value
            time.sleep(0.5)
                # Aquí llamas a tu función de clicks, por ejemplo:
            if contador==0:
                print("Celda vacía, ejecutando clicks...")  
                time.sleep(0.5)
                moveToClickAndWait(coord_desmarcar[0],coord_desmarcar[1],0,0)
                #algunos archivos de la cnbv no es necesario marcar y desmarcar la casilal de bancos para selccioanr todos basta con un solo marcado
                #agregar o quitar la siguiente linea segun sea el caso
                moveToClickAndWait(coord_banco[0],coord_banco[1],0,0)
                moveToClickAndWait(coord_banco[0],coord_banco[1],2,0)
                moveToClickAndWait(coord_abajo[0],coord_abajo[1],0,6)
                moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                
            elif contador<=max_iteraciones-3:
                print(f"Celda cambió: {valor_actual}")
                # Continuar con siguiente acción o rango
                traer_tabla(ws_cnbv)
                marcarSiguienteFecha(20-contador)
            else:
                traer_tabla(ws_cnbv)
                moveToClickAndWait(coord_fecha[0],coord_fecha[1],0,0)
                moveToClickAndWait(coord_penultima[0],coord_penultima[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                traer_tabla(ws_cnbv)
                moveToClickAndWait(coord_penultima[0],coord_penultima[1],0,0)
                moveToClickAndWait(coord_ultima[0],coord_ultima[1],0,0)
                moveToClickAndWait(coord_obtener[0],coord_obtener[1],0,0)
                traer_tabla(ws_cnbv)
                break
            contador += 1
    
    wb_cnbv.Close(SaveChanges=False)        
    excel.Quit()
####################Ruta de prueba##########################
x=gv.ARCHIVOS_DATA_RAW[3]["ruta"]


def validar_ruta(ruta:Path)->Path:
    if not ruta.exists() or not ruta.is_file():
        raise FileNotFoundError(
                    f"No se encontró el archivo: {ruta.name}\n"
                    f"Se buscó en la ruta: {ruta.parent}\n"
                    f"Por favor, verifica la carpeta \"Data\""
                )
    return ruta

def extraer_datos(celda:Any,rutadestino:Path,ventanaexcel)->None:
    rango=celda.CurrentRegion()
    rutadellibrocontrol=validar_ruta(gv.Paths.output/"LibroControl.xlsx")
    nombreventana=rutadellibrocontrol.stem
    try:
        libroabierto=ventanaexcel.Workbooks(nombreventana)
        libroabierto.Close(SaveChanges=False)
        print("Se cierra la ventana previa del LibroControl si es que esta abierto para evitar inyectar informacion erronea")
    except:
        pass


    librocontrol=ventanaexcel.Workbooks.Open(str(rutadellibrocontrol))
    librocontrol.Activate()

    filas=len(rango)
    columnas=len(rango[0])
    hoja=librocontrol.Sheets(1)

    print(type(rango))
    celdas_destino=hoja.Range(hoja.Cells(1, 1),hoja.Cells(filas, columnas))
    celdas_destino.Value=rango
    try:
        ventanaexcel.DisplayAlerts = False
        librocontrol.SaveAs(str(rutadestino))
        librocontrol.Close(SaveChanges=False)
    except Exception as e:
        print(f"Error en el metodo SaveAs , no se pudo guardar {type(e).__name__}: {e}\n")
    finally:
            if librocontrol is not None:
                try:
                    ventanaexcel.DisplayAlerts = False 
                    librocontrol.Close(SaveChanges=False)
                except:
                    ventanaexcel.DisplayAlerts = True
                    pass     
    return None

def string_eureka_fechas(libro:Any)->tuple:
    primerhoja=libro.Sheets(1)
    primerhoja.Activate()
    # buscar_celda=primerhoja.Cells.Find(What="Banco")
    # if buscar_celda:
    #     print(f"Encontrado en: {buscar_celda.Address}")
    # print(buscar_celda)
    buscar_1=primerhoja.Range("B1").End(-4121).End(-4121)

    if not primerhoja.Range("B1").End(-4121).Value=="Notas":
        raise Exception("\nCambio el formato de la CNBV actualizar saltos de linea y espaciados de nuevo para encontrar la tabla, se debe modificar los pasos de la libreria win32com.client.")

    a=buscar_1.CurrentRegion
    a.Activate()
    tupla_datos=a.Value

    if not len(tupla_datos[0])>=2:
        raise Exception("\nEs muy probable que cambio la formato de la CNBV pues no se esta alcanzando la tabla de datos para su extraccion.")
    try:
    #celda=primerhoja.Cell
        str_fecha=str(int(tupla_datos[0][2]))

        print("eureka")        
        print("Formato estandar detectado no se necesito aplicar logica adicional")
    except (IndexError, ValueError, TypeError):
        str_fecha = str(int(tupla_datos[1][2]))
        print("Formato alternativo detectado")

    if not re.match(con.patrones["fecha"],str_fecha):
        raise Exception("\nEl valor para buscar la fecha y nombrar las carpetas no se cumple. Validar la celda de donde se esta extrayendo la fecha.")
        
    str_fecha=str_fecha
    celda_tabla=buscar_1

    return str_fecha,celda_tabla  



def abrir_excel_por_ruta(rutaexcel:Path)->tuple:
    #Validamos que la ruta exista
    ruta=validar_ruta(rutaexcel)

    ventanaex = win32.Dispatch("Excel.Application")
    hdwnd_excel = ventanaex.Hwnd
    ventanaex.Visible=True



    try:
        libro=ventanaex.Workbooks.Open(ruta)
        win32gui.ShowWindow(hdwnd_excel, win32con.SW_MAXIMIZE)
        win32gui.SetForegroundWindow(hdwnd_excel)

    #tener cuidado de no abrir un libro antes por que si no falla traer al frente el archivo de consultas

    except FileNotFoundError:
        print("El archivo no existe en esa ruta")
    except Exception as e:
        print(f"Ocurrió un error inesperado con la libreria win32com.client {type(e).__name__}: {e}\n")
    print("Excel abierto y traido al frente")
    #libro.Close(False)
    #ventanaex.Quit()
    return libro,ventanaex


def revisar_carpeta_output_y_crear_carpeta(archivoecxel: Path)-> Path:
    ruta=validar_ruta(archivoecxel)
    gv.RUTA_OUTPUT.mkdir(parents=True, exist_ok=True)
    rutaNueva=gv.RUTA_OUTPUT / busqueda_patron_para_nombrar_carpetas(ruta.name)
    rutaNueva.mkdir(parents=True, exist_ok=True)
    return rutaNueva

def one_click(numerodeclicks,coordenada)->None:
    c=0
    while (numerodeclicks)>=c:
        pyautogui.click(coordenada)
        c+=1
    return None

def final_bimestres_y_seleccionar_todos_los_bancos(numerodeclicks)->None:
    one_click(numerodeclicks-3,coord_abajo)
    one_click(2,coord_banco)
    #bloquepara obtener informacion aqui es un punto critico de fallos
    one_click(1,coord_obtener)
#Revisa el archivo para saber cuantos bimestres contiene, y genera un lista de los bimestres contenidos


def lista_verdad(fecha:str)->list:
    
    con.patron.revisar_patron(con.patron.FECHAS,fecha)
    yearbase=2011
    try:
        añosultimobimestre=int(fecha[:4])
        añostranscurridos=añosultimobimestre-yearbase
        bimestreactual=int(fecha[-2:])
    except ValueError as e:
        print(f"Formato del string fecha incorrecto {e}")
    bimestresdel2011=4
    bimestresenarchivo=6*(añostranscurridos-1)+(bimestresdel2011+int(bimestreactual/2))
    print(f"bimestres en el archivo:{bimestresenarchivo}")
    #para resolver el seguimiento de fechas se solcuiona con una transformacion lineal que planteo su servidor y tener una lista de todos los bimestres 
    #que estan contenidos en el archivo excel al momento de abrirlo
    y=[]
    yearconstante=201104
    c=0
    for i in range(1,bimestresenarchivo+1):

        yearconstante+=2    
        if (yearconstante-201100-100*c)==14:
            c+=1 
            yearconstante+=100
            yearconstante-=12

        y.append(yearconstante)


    return y



#orquesta(x)


#implementacion de clases para el manejo de rutas y archivos creados
class rutaCarpetas:
    def __init__(self,rutasInsts:con.Paths,archivoExcel:Path)->None:
        self.rutasInsts=rutasInsts
        self.nombreArchivExcel=archivoExcel.stem
        self.nombreSerie=con.patron.busqueda_patron_para_nombrar_carpetas(self.nombreArchivExcel)
        self.carpetaDestino=self.rutasInsts.output / self.nombreSerie
        self.carpetaDestino.mkdir(parents=True, exist_ok=True)


    @property
    def revisar_archivos_en_carpeta(self)->list:
        listaarchivos=[]
        for archivos in self.carpetaDestino.iterdir():
            if archivos.is_file():
                listaarchivos.append(archivos.stem)
        return listaarchivos

R=rutaCarpetas(rutasInstaciadas,gv.ARCHIVOS_DATA_RAW[0]["ruta"])
print(R.revisar_archivos_en_carpeta[0])
class carajo:
    def __init__(self,hojaexcel:Any,mapa:rutaCarpetas)->None: 
        self.hojaExcel=hojaexcel
        self.fecha,self.celda_tabla=string_eureka_fechas(self.hojaExcel)
        #La lista de los bimestres contenidos en el archivo excel
        self.listaFechas=lista_verdad(self.fecha)
        self.archivosCreados=mapa.revisar_archivos_en_carpeta
    @property    
    def pendientes_de_consulta(self)->list:

        pass
        


#buscar_fechas_en_el_Activex_y_extraer(x,y)
class boton:
    def __init__(self,fecha:int,posicion:int,estadoInicial:bool|None=None):
        self.fecha=fecha
        self.posicion=posicion
        self.estado=False
    @property
    def estado_actual(self)->bool:
        return self.estado
    
    def cambiar_estado(self)->None:
        self.estado=not self.estado

class coordenadaSimple:
    def __init__(self,coordenada:tuple,intervalo:int)->None:
        self.coordenada=coordenada
        self.intervalo=intervalo 

    def click_en_coordenada(self,boton:boton|None=None)->None:
        pyautogui.click(self.coordenada,interval=self.intervalo)
        if boton:
            boton.cambiar_estado()

#Esta clase servira para saber si la cordenada dentro del checkbox esta activa o no, y nos porporcionara un boton mental
class pointCheckboxManager(coordenadaSimple):
    def __init__(self,config_visual:dict,radio:Any|None=None)->None:
        self.radio=radio
        self._despachador=config_visual

        super().__init__(self._despachador["fecha_top"],self._despachador["tiempo_espera_entre_clicks"])
        
    def click_auxiliar_en_coordenada(self,punto:tuple)->None:
        pyautogui.click(punto,interval=self._despachador["tiempo_espera_entre_clicks"])   

    def siguiente_cordenada(self,pasos:int):
        pyautogui.click(self._despachador["abajo"],clicks=abs(pasos),interval=self._despachador["tiempo_espera_entre_clicks"])
        self.radio._propiocepcion+=pasos

    def  anterior_cordenada(self,pasos:int):
        pyautogui.click(self._despachador["arriba"],clicks=abs(pasos),interval=self._despachador["tiempo_espera_entre_clicks"])
        self.radio._propiocepcion-=pasos

    def obtener(self,casoparticular:bool=False)->bool:
        pyautogui.click(self._despachador["obtener"])
        time.sleep(CONFIG_COORDS["tiempo_espera_consulta"])

class automataConsultas:
    def __init__(self,punto_actual:pointCheckboxManager,obervador:carajo)->None:
        self.hdwm_excel=None
        self.observador=obervador
        #iniciamos la matriz de fechas que hay dentro del archivo excel que se esta consultando
        self.matrizFechas=observador.listaFechas
        #con las fechas creamos botones con todos inicializados en falso, osea desmarcados y solo el primero queda marcado tal y como esta el estandar
        #del archivo excel, la fecha mas reciente esta marcada y consultada
        self.matrizBotones:dict[int,boton]={}
        self._navegante=punto_actual
        #le pasamos la radio para que pueda actualizar su propiocepcion
        self._navegante.radio= self
        self._propiocepcion = 0
        #obtenemos la hoja excel del observador
        self.hojaExcel=self.observador.hojaExcel

        self.propiocepcion_actual
        self.matrizFechas.reverse()
        try:
            self.crear_botones()
            # Como la primer fecha siemproe esta marcada el primer boton queda activo tal como los archivos
            #accedemeos manualmenete al dictioanro con el metodo iter para ir uno a uno y para el primero encapsualmos en un next
            # se implemnta nueva busqueda por la hoja excel en vez de inicializar las variables, no descarto la inicializacion de variables de momento se quitara
            primer_boton = next(iter(self.matrizBotones.values()))
            primer_boton.estado = True
        except Exception as e:
            print(f"Error critico no se pudieron crear los botones para guiar al automata, revisar la propiedad \"self.matrizFechas\"\n{e}")
            self.matrizBotones={}

        #con lo que hacemos la propiocepcion es con la fecha que se obtiene al iniciar el automata
        self.ojos=string_eureka_fechas(self.hojaExcel)[0]
    
    def crear_botones(self):


        for k, fechas in enumerate(self.matrizFechas):
            x=boton(fechas,k)
            self.matrizBotones[fechas]=x
        

    def avanzar_a_fecha(self,pasos:int,freno:int|None=None)->None:

        if pasos == 0:
            return
        
        if freno and abs(pasos)>=freno:
            avance=freno
        else:
            avance=pasos

        if pasos>0:
            self._navegante.siguiente_cordenada(avance)
        elif pasos<0:
            self._navegante.anterior_cordenada(avance)

        print("propiocepcion",self._propiocepcion)

        if freno and abs(pasos)>=freno:
            self._navegante.click_en_coordenada()

        self.avanzar_a_fecha(pasos-avance,freno)

    
    def ir_apagar_boton(self,boton:boton):

        if boton.estado:
            self.avanzar_a_fecha(boton.posicion-self._propiocepcion)
            self._navegante.click_en_coordenada(boton)
        

    def ir_apagar_todos_los_botones_encendidos(self)->None:
        for boton in self.matrizBotones.values():
            if boton.estado:
                self.ir_apagar_boton(boton)

    def invertir_estado_varios_botones(self,cambios:list)->None:
        for botonFecha in cambios:
            self.matrizBotones[botonFecha].cambiar_estado()

    @property            
    def botones_encendidos(self)->list:
        listaFechas=[]
        for boton in self.matrizBotones.values():
            if boton.estado:
                listaFechas.append(boton.fecha)
        return listaFechas

    @property
    def propiocepcion_actual(self)->int:
        print (self._propiocepcion)
        return self._propiocepcion

    def homming(self)->None:
        activar_ventana_al_frente(self.hdwm_excel)
        self.avanzar_a_fecha(-self._propiocepcion)
        self._navegante.click_en_coordenada()

    def obtener_fecha_actual(self,fechaRevision:int,casoespecial=False)->bool:


        
        try:
            
            print(f"fecha actual obtenida: {fechaRevision}, coincide con la propiacepcion {self.matrizFechas[self._propiocepcion]}")
            print("fecha coincide con la actual, obtener")
            if not fechaRevision==self.matrizFechas[self._propiocepcion]: 
                print("Implemnetando nueva logica")

                return True
            else: 

                return False
            

        except IndexError as e:
            print(f"IndexError en obtener fecha actual, probablemente la propiocepcion se salio de rango No es un error critico continua {e}")
            return False




    @property
    def listadefechas(self)->None:
        print("lista de fechas del observador",self.observador.listaFechas)    
        return None
    @property
    def limpiar(self)->None:
        print("lista de fechas del observador",self.observador.listaFechas)    
        return None

    def recorrer_pendientes(self,lisapendientes:list)->None:
        print(self.matrizFechas)
        print("lista de pendientes a consultar",lisapendientes)


        cicloRoto=False
        posicion=0
        freno=None

        for k,indice in enumerate(lisapendientes):

            if k==0 and self.matrizFechas[k] not in [self.matrizFechas[i] for i in lisapendientes]:
                self.ir_apagar_boton(self.matrizBotones[self.matrizFechas[k]])

            if indice==7:
                print("para para el debugger")


            self.avanzar_a_fecha(indice-self._propiocepcion)
            fechaEnBucle=self.matrizFechas[indice]
            
            #validamos que el boton no es te seleccionado  
            if not self.matrizBotones[fechaEnBucle].estado:
                self._navegante.click_en_coordenada(boton=self.matrizBotones[fechaEnBucle])


            self._navegante.obtener()
            fechaEnHoja=int(string_eureka_fechas(self.hojaExcel)[0])

            if self.obtener_fecha_actual(fechaEnHoja):  
                cicloRoto=True


                ajustePropiocepcion=self.matrizBotones[fechaEnHoja].posicion-self._propiocepcion
                self._propiocepcion+=ajustePropiocepcion

                #cuando se desajusta para encontrar el boton que se encendio por error, es la diferencia entre el anterior indice menos el indice actual 
                #de la lista de pendientes en la depuracion de una lista ejemplo [3,7,11], pasos=7-3=4 que es el borton que se encendio por error 
                pasos=lisapendientes[k]-lisapendientes[k-1]
                self.invertir_estado_varios_botones([fechaEnHoja,fechaEnBucle,self.matrizFechas[lisapendientes[k-1]],self.matrizFechas[pasos]])
                self.ir_apagar_todos_los_botones_encendidos()

                print("entro al caso particular")
                break

            self.ir_apagar_boton(self.matrizBotones[fechaEnHoja])
            posicion+=1   


 

            print("propiocepcion despues de avanzar",self._propiocepcion)



        print("fecha de cambio",fechaEnHoja)

        if cicloRoto:
            fechaQueRompeElBucle=fechaEnHoja
            ajuste=indice
            for indice in lisapendientes[posicion:]:

                fechaEnBucle=self.matrizFechas[indice]
                #Pruebas para ajustar la propiocepcion con ayuda del debugger , simular los clciks manualmente mientras se testea
                #print(indice)
                #print(self.observador.listaFechas[-self._propiocepcion])
                if indice==11:
                    print("para revisar con el debugger")
                # 201406, 201310, 201302,201206,
                self.avanzar_a_fecha(indice-self._propiocepcion)
                self._navegante.click_en_coordenada(boton=self.matrizBotones[fechaEnBucle])
                
                #print(self.observador.listaFechas[-self._propiocepcion])
                if  not self.matrizFechas[self._propiocepcion] in [201108,201106]:

                    self._navegante.obtener()
                    fechaEnHoja=int(string_eureka_fechas(self.hojaExcel)[0])
                    self.obtener_fecha_actual(fechaEnHoja) 
                    self._propiocepcion-=self.matrizBotones[fechaQueRompeElBucle].posicion+self._propiocepcion
                    


                    self.ir_apagar_boton(self.matrizBotones[fechaEnHoja])


 
                elif self.matrizFechas[self._propiocepcion]==201108:
                    self._navegante.click_en_coordenada()
                    self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["penultima"])
                    self._navegante.obtener()
                    continue

                elif self.matrizFechas[self._propiocepcion]==201106:
                    self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["penultima"])
                    self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["fecha_bottom"])             
                    self._navegante.obtener()
                    break 


        #self.homming()
        print("Proceso de consultas finalizado")




        # if  salio_del_bucle:
        #     self.avanzar_a_fecha(self._propiocepcion)
        #     self._navegante.click_en_coordenada()
        #     self.obtener_fecha_caso_particular

        #     print(self.propiocepcion_actual())       
# IMPLEMENTACION DE NUEVA LOGICA PRA CUANDO EL ARCHIVO REGRESA A UNA CASILLA EN ESPECIFICO DEL CUADRO ACTIVEX


class excelOpenManager:

    def __init__(self,ruta):
        self.ventana=win32.Dispatch("Excel.Application")
        self.hdwn=self.ventana.Hwnd
        self.libro
        pass



time.sleep(2)   
casilla1=pointCheckboxManager(CONFIG_COORDS)
mapa=rutaCarpetas(rutasInstaciadas,x)

lista=[202302,202304,202306,202308,202310,202312,202402,202404,202406,202408,202410,202412,202502,202504,202506]
lista.reverse()

lista_pendientes=[202302,202408,202410]
indices = []
c=0
for i in lista:
    c+=1
    for k in lista_pendientes:
        if i==k:

            indices.append(c)
            break
print(indices)

libro,ventana=abrir_excel_por_ruta(x)

observador=carajo(libro,mapa)

hoja1=automataConsultas(casilla1,observador)
hoja1.hdwm_excel=ventana.Hwnd

hoja1.recorrer_pendientes(range(5))


# print(hoja1.matrizFechas)

#################################### 
#Implementacion de decoradores
def mi_primer_decorador(fechaBuscada:str):

    def decorador_real(funcion):

            def envoltura(*args, **kwargs):
                pass




#print(type(gv.ARCHIVOS_DATA_RAW[0]["ruta"]))
#print(type(Path().mkdir))
# 1. Definimos la ruta usando Pathlib (como lo hablamos hoy)
# ventanaex = win32.Dispatch("Excel.Application")
# libro=ventanaex.Workbooks.Open(x).Worksheets(1).Cells
# ruta_txt = Path("metodos_excel_hoja.txt")

# 2. Obtenemos la lista de atributos y la limpiamos
# Filtramos los que empiezan con "_" porque suelen ser internos de Python
# atributos = [attr for attr in dir(libro) if not attr.startswith("_")]

# 3. Guardamos en el archivo
# with ruta_txt.open("w", encoding="utf-8") as f:
#     f.write(f"MÉTODOS Y PROPIEDADES DISPONIBLES PARA: {type(libro)}\n")
#     f.write("="*50 + "\n")
#     for attr in atributos:
#         f.write(f"{attr}\n")

# print(f"✅ ¡Listo! Se han guardado {len(atributos)} métodos en {ruta_txt.absolute()}")

        # for k, indice in enumerate(lisapendientes[posicion:]):


        #     Pruebas para ajustar la propiocepcion con ayuda del debugger , simular los clciks manualmente mientras se testea
        #     print(indice)
        #     print(self.observador.listaFechas[-self._propiocepcion])
        #     if indice==4:

        #         print("entro al caso especial 4")

        #     self._propiocepcion-=ajuste
            
        #     self.avanzar_a_fecha(indice-ajuste+1-self._propiocepcion)
        #     print(self.observador.listaFechas[-self._propiocepcion])
        #     if  not self.matrizFechas[self._propiocepcion] in [201108,201106]:
        #         self._navegante.click_en_coordenada()

        #         self._propiocepcion+=ajuste
        #         self._propiocepcion+=ajuste-2
        #         self.obtener_fecha_actual(casoespecial=True)
        #         self._propiocepcion-=ajuste-2




        #     self.avanzar_a_fecha(ajuste)
        #     ajuste+=1

        #     if self.matrizFechas[self._propiocepcion]==201108:
        #         self._navegante.click_en_coordenada()
        #         self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["penultima"])
        #         self.obtener_fecha_actual(casoespecial=True) 
        #         continue

        #     if self.matrizFechas[self._propiocepcion]==201106:
        #         self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["penultima"])
        #         self._navegante.click_auxiliar_en_coordenada(self._navegante._despachador["fecha_bottom"])             
        #         self.obtener_fecha_actual(casoespecial=True)
        #         break 

        #     self._navegante.click_en_coordenada()